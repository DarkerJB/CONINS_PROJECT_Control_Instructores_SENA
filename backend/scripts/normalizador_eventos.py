import re, sys, unicodedata, datetime
import openpyxl
from openpyxl import Workbook

SEED = "/sessions/sleepy-nifty-lovelace/mnt/CONINS_PROJECT_Control_Instructores_SENA/backend/seed_data.sql"
RAW  = "/sessions/sleepy-nifty-lovelace/mnt/CONINS_PROJECT_Control_Instructores_SENA/docs/ConIns_ERS_Y_Relevantes/Eventos - ADSO - Agosto - Enviar.xlsx"
OUT  = "/sessions/sleepy-nifty-lovelace/mnt/CONINS_PROJECT_Control_Instructores_SENA/docs/ConIns_ERS_Y_Relevantes/Eventos_ADSO_Agosto_NORMALIZADO.xlsx"
PROGRAMA = "228118"  # ADSO

def norm(s):
    if s is None: return ""
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.lower()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

seed = open(SEED, encoding="utf-8").read()

# --- mapa competencia: nombre_norm -> codigo (+ lista para fuzzy) ---
comp_block = re.search(r"INSERT IGNORE INTO competencias \(id, nombre, codigo, programa_id\) VALUES(.*?);", seed, re.S)
comps = []
if comp_block:
    for m in re.finditer(r"\(\s*\d+\s*,\s*'([^']*)'\s*,\s*'([^']*)'\s*,\s*\d+\s*\)", comp_block.group(1)):
        comps.append((norm(m.group(1)), m.group(2), m.group(1)))

def match_competencia(text):
    t = set(norm(text).split())
    if not t: return None, 0
    best, bestc = 0, None
    for nn, cod, orig in comps:
        s = set(nn.split())
        if not s: continue
        j = len(t & s) / len(t | s)
        if j > best: best, bestc = j, cod
    return (bestc, round(best,2)) if best >= 0.35 else (None, round(best,2))

# --- mapa instructor: nombre_norm -> email ---
usr = {}
for m in re.finditer(r"\(\s*\d+\s*,\s*'([^']*)'\s*,\s*'([^']*@[^']*)'", seed):
    usr[norm(m.group(1))] = m.group(2)
def match_email(name):
    n = norm(name)
    if n in usr: return usr[n]
    tn = set(n.split())
    for k,v in usr.items():
        ks = set(k.split())
        if tn and (tn <= ks or ks <= tn) and len(tn & ks) >= 2:
            return v
    return None

JMAP = {"MANANA":"manana","MAÑANA":"manana","TARDE":"mixta","NOCHE":"noche"}
JHORA = {"manana":("06:00","12:00"),"mixta":("12:00","18:00"),"noche":("18:00","22:00")}
DIAS = {0:1,1:2,2:3,3:4,4:5,5:6,6:7}

def parse_ficha(s):
    s = str(s or "")
    fic = re.search(r"\b(\d{7})\b", s)
    jor = None
    for k,v in JMAP.items():
        if k in s.upper(): jor = v; break
    hor = re.search(r"(\d{1,2}:\d{2})\s*A\s*(\d{1,2}:\d{2})", s)
    hi = hf = None
    if hor:
        hi, hf = hor.group(1), hor.group(2)
        def ok(x):
            try:
                h,mn = map(int,x.split(":")); return 0<=h<=23 and 0<=mn<=59
            except: return False
        if not (ok(hi) and ok(hf)): hi=hf=None
    return (fic.group(1) if fic else None), jor, hi, hf

def parse_fechas(cell):
    out=[]
    for line in str(cell or "").split("\n"):
        m = re.match(r"\s*(\d{2})/(\d{2})/(\d{4})\s*$", line.strip())
        if m:
            d,mo,y = map(int, m.groups())
            try: out.append(datetime.date(y,mo,d))
            except: pass
    return out

def parse_ambiente(v):
    s = str(v or "").strip()
    if not s: return ""
    if re.fullmatch(r"\d+", s): return f"Ambiente {s}"
    if "estrella" in s.lower(): return "Aula Ambiental La Estrella"
    return " ".join(w.capitalize() for w in re.sub(r"\s+"," ",s.replace("\n"," ")).split())

wb = openpyxl.load_workbook(RAW, data_only=True)
instructores={}; grupos={}; asignaciones={}; horarios=[]; problemas=[]
TRANSV_SHEETS = {"bilinguismo","transversales"}

for ws in wb.worksheets:
    tipo_area = "transversal" if ws.title.strip().lower() in TRANSV_SHEETS else "tecnica"
    for r in ws.iter_rows(values_only=True):
        inst = r[1] if len(r)>1 else None
        if not isinstance(inst,str): continue
        u = inst.strip()
        if not u or u.upper()=="INSTRUCTORES" or "EVENTOS" in u.upper(): continue
        name = u.split(" - ")[0].strip()
        email = match_email(name)
        comp_txt = r[3] if len(r)>3 else None
        rap_txt  = r[4] if len(r)>4 else None
        fecha_c  = r[5] if len(r)>5 else None
        amb_c    = r[7] if len(r)>7 else None
        fic_c    = r[8] if len(r)>8 else None
        ficha, jor, hi, hf = parse_ficha(fic_c)
        if jor and not (hi and hf): hi,hf = JHORA.get(jor,(None,None))
        cod, score = match_competencia(comp_txt)
        amb = parse_ambiente(amb_c)
        fechas = parse_fechas(fecha_c)
        if not email: problemas.append((ws.title,name,"instructor sin email en catalogo")); continue
        if not ficha: problemas.append((ws.title,name,"ficha no detectada")); continue
        if not cod:   problemas.append((ws.title,name,f"competencia sin match ({str(comp_txt)[:40]}...)")); continue
        instructores[email]=(name,tipo_area,cod)
        grupos.setdefault(ficha,(jor or "", amb))
        asignaciones[(email,ficha,cod)]=True
        for d in fechas:
            lunes = d - datetime.timedelta(days=d.weekday())
            horarios.append((email,ficha,cod,"",amb,DIAS[d.weekday()],hi or "",hf or "",jor or "",lunes.isoformat()))

# --- escribir plantilla ---
out = Workbook(); out.remove(out.active)
sh = out.create_sheet("Instructores"); sh.append(["nombre","email","tipo_area","codigos_competencia"])
byemail={}
for (email,fic,cod) in asignaciones: byemail.setdefault(email,set()).add(cod)
for email,(name,ta,_) in instructores.items():
    sh.append([name,email,ta,";".join(sorted(byemail.get(email,[])))])
sg = out.create_sheet("Grupos"); sg.append(["numero_grupo","codigo_programa","jornada","ambiente"])
for fic,(jor,amb) in sorted(grupos.items()): sg.append([fic,PROGRAMA,jor,amb])
sa = out.create_sheet("Asignaciones"); sa.append(["instructor_email","numero_grupo","codigos_competencia"])
seen=set()
for (email,fic,cod) in asignaciones:
    k=(email,fic)
    if k in seen: continue
    seen.add(k); sa.append([email,fic,";".join(sorted(byemail[email]))])
sho = out.create_sheet("Horarios"); sho.append(["instructor_email","numero_grupo","codigo_competencia","codigo_rap","ambiente","dia_semana","hora_inicio","hora_fin","jornada","semana"])
for h in horarios: sho.append(list(h))
out.save(OUT)

print("INSTRUCTORES:",len(instructores),"| GRUPOS:",len(grupos),"| ASIGNACIONES:",len(seen),"| HORARIOS:",len(horarios))
print("PROBLEMAS (filas no mapeadas):",len(problemas))
from collections import Counter
for msg,c in Counter(p[2].split('(')[0].strip() for p in problemas).most_common(): print(f"  {c}x {msg}")
print("OUT:",OUT)
